from openpyxl import load_workbook, Workbook
import csv

class Reader():
    def __init__(self,filename):
        self.filename = filename[0]

    def read(self): pass


class ExcelReader(Reader):
    def __init__(self, filename):
        super(ExcelReader, self).__init__(filename)

    def read(self, sheet_name="Sheet1"):
        workbook = load_workbook(self.filename)
        data = workbook[sheet_name]
        rows = data.max_row

        # 打印excel总行数
        print("rows total: %s." % rows)

        # 初始化商品url列表，处理后返回
        goods_url_list = []

        for r in range(2, data.max_row + 1):
            goods_url_list.append(data.cell(r, 1).value)

        return goods_url_list

class TxtReader(Reader):
    def __init__(self, filename):
        super(TxtReader, self).__init__(filename)

    def read(self):
        goods_url_list = []
        with open(self.filename, "r") as f:
            for url in f.readlines():
                goods_url_list.append(url.strip())

        return goods_url_list


class CsvReader(Reader):
    def __init__(self, filename):
        super(CsvReader, self).__init__(filename)
        
    def read(self):
        goods_url_list = []
        with open(self.filename, "r") as f:
            csvreader = csv.reader(f)
            for url in csvreader:
                goods_url_list.append(url[0])

        return goods_url_list


READER_CLASS = {
    ".txt" : TxtReader.__name__,
    ".xlsx" : ExcelReader.__name__,
    ".csv": CsvReader.__name__
}
