from openpyxl import load_workbook, Workbook
import os

class Writer():
    def __init__(self, filename):
        self.filename = filename[0]

    def write(self, data): pass

    def save(self): pass

class ExcelWriter(Writer):
    def __init__(self, filename):
        super(ExcelWriter, self).__init__(filename)
        self.workbook = Workbook()
        self.row = 1

    def write(self, data):
        sheet = self.workbook.active
        sheet.title = "Sheet1"
        print(data)
        excel_column_index = list(range(2, len(data) + 2))
        print(excel_column_index)
        goods = list(zip(excel_column_index, data))
        for index, goods_key in goods:
            sheet.cell(self.row, index).value = data[goods_key]
        self.row += 1

    def save(self):
        self.workbook.save(self.filename)

class TxtWriter(Writer):
    def __init__(self, filename):
        super(TxtWriter, self).__init__(filename)
        

    def write(self, data): 
        self.f = open(file=self.filename,mode="a+")

        i = 0
        for key in data:
            value = data[key]
            self.f.write("%s:%s, " % (key , value))
            i += 1

            if i == len(data):
                self.f.write("\n")
            

    def save(self): 
        self.f.close()

class EmailSender(Writer):
    # TODO Change from target file to target mailbox
    def __init__(self, filename):
        super(EmailSender, self).__init__(filename)
        self.content = ''

    def write(self,data):
        i = 0
        for key in data:
            value = data[key]
            self.content += "%s:%s, " % (key , value)
            i += 1

            if i == len(data):
                self.content += "\n"

                
    def send(self):
        # TODO Send self.content to corresponding mailbox
        pass



WRITER_CLASS = {
    ".txt" : TxtWriter.__name__,
    ".xlsx" : ExcelWriter.__name__,
    ".email": EmailSender.__name__  # TODO .email <- It's temporary
}